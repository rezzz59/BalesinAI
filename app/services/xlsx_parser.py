"""XLSX parser — wraps openpyxl and normalizes rows to canonical columns.

Supports either a single sheet containing both FAQ and catalog-like columns,
or separate "FAQ"/"Katalog" sheets. Reuses the canonical column maps from
app.services.sheets so uploaded data behaves identically to Google Sheets.
"""
import io
import logging

from openpyxl import load_workbook

from app.services.sheets import CATALOG_COL_MAP, FAQ_COL_MAP, _normalize_header

logger = logging.getLogger(__name__)


class XlsxParseError(Exception):
    """Raised when the uploaded file has no usable FAQ/catalog data."""


def _match_column(headers: list[str], col_map: dict[str, list[str]]) -> dict[str, int]:
    """Map canonical keys (e.g. 'pertanyaan') -> column index, by header alias."""
    result: dict[str, int] = {}
    norm_headers = [(_normalize_header(h), i) for i, h in enumerate(headers)]
    for canonical, aliases in col_map.items():
        for alias in aliases:
            want = _normalize_header(alias)
            for norm, idx in norm_headers:
                if norm == want:
                    result.setdefault(canonical, idx)
                    break
    return result


def shape_faq_rows(rows: list[dict]) -> list[dict]:
    return [
        {"pertanyaan": r.get("pertanyaan", ""), "jawaban": r.get("jawaban", "")}
        for r in rows if (r.get("pertanyaan") or "").strip() and (r.get("jawaban") or "").strip()
    ]


def shape_catalog_rows(rows: list[dict]) -> list[dict]:
    return [
        {
            "nama_produk": r.get("nama_produk", ""),
            "harga": r.get("harga", ""),
            "ready": r.get("ready", ""),
            "deskripsi": r.get("deskripsi", ""),
        }
        for r in rows if (r.get("nama_produk") or "").strip()
    ]


def parse_workbook(content: bytes) -> dict:
    """Parse an uploaded XLSX. Returns {'faq': [...], 'catalog': [...]} dicts.

    Recognizes: one sheet with FAQ columns, one sheet with catalog columns, or
    a single sheet that contains both. Raises XlsxParseError if nothing usable.
    """
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        raise XlsxParseError(f"File tidak bisa dibaca: {e}")

    faq_rows: list[dict] = []
    catalog_rows: list[dict] = []

    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        header = None
        body: list[list] = []
        for row in it:
            vals = list(row)
            if header is None:
                if not any(c is not None and str(c).strip() for c in vals):
                    continue
                header = [str(c).strip() if c is not None else "" for c in vals]
                continue
            body.append(vals)

        if not header:
            continue

        faq_cols = _match_column(header, FAQ_COL_MAP)
        cat_cols = _match_column(header, CATALOG_COL_MAP)

        def pick(row, col_idx: int):
            if col_idx is None or col_idx >= len(row):
                return ""
            return str(row[col_idx]) if row[col_idx] is not None else ""

        if "pertanyaan" in faq_cols and "jawaban" in faq_cols:
            for row in body:
                q = pick(row, faq_cols["pertanyaan"]).strip()
                a = pick(row, faq_cols["jawaban"]).strip()
                if q and a:
                    faq_rows.append({"pertanyaan": q, "jawaban": a})

        if "nama_produk" in cat_cols:
            for row in body:
                name = pick(row, cat_cols["nama_produk"]).strip()
                if not name:
                    continue
                catalog_rows.append({
                    "nama_produk": name,
                    "harga": pick(row, cat_cols["harga"]) if "harga" in cat_cols else "",
                    "ready": pick(row, cat_cols["ready"]) if "ready" in cat_cols else "",
                    "deskripsi": pick(row, cat_cols["deskripsi"]) if "deskripsi" in cat_cols else "",
                })

    wb.close()
    if not faq_rows and not catalog_rows:
        raise XlsxParseError(
            "Tidak ada data FAQ (kolom pertanyaan/jawaban) atau katalog "
            "(kolom nama produk) yang dikenali di file ini."
        )
    return {"faq": faq_rows, "catalog": catalog_rows}