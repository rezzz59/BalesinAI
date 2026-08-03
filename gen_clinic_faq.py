import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Data FAQ untuk klinik Indonesia
# Format: (pertanyaan, jawaban, kategori)
faq_data = [
    # ===== 1. INFORMASI UMUM (1-10) =====
    ("Apa nama klinik ini?", "Nama klinik kami adalah Klinik Sehat Bersama.", "Informasi Umum"),
    ("Di mana lokasi klinik?", "Klinik kami berlokasi di Jl. Sudirman No. 123, Jakarta Selatan.", "Informasi Umum"),
    ("Jam operasional klinik?", "Kami buka setiap hari Senin-Sabtu pukul 08.00-21.00 WIB, dan Minggu pukul 09.00-17.00 WIB.", "Informasi Umum"),
    ("Apakah klinik buka hari libur?", "Klinik tetap buka pada hari libur besar dengan jam operasional terbatas (09.00-15.00).", "Informasi Umum"),
    ("Apakah klinik 24 jam?", "Tidak, klinik kami buka dari pagi hingga malam sesuai jam operasional.", "Informasi Umum"),
    ("Berapa nomor telepon klinik?", "Nomor telepon klinik: (021) 1234-5678.", "Informasi Umum"),
    ("Apakah ada WhatsApp klinik?", "Ya, WhatsApp klinik: 0812-3456-7890.", "Informasi Umum"),
    ("Apakah ada aplikasi mobile?", "Belum, tapi Anda bisa booking via WhatsApp atau datang langsung.", "Informasi Umum"),
    ("Apakah klinik ini rumah sakit?", "Kami adalah klinik praktik dokter, bukan rumah sakit.", "Informasi Umum"),
    ("Apakah klinik ini termasuk BPJS?", "Ya, kami bekerja sama dengan BPJS Kesehatan dan beberapa asuransi swasta.", "Informasi Umum"),

    # ===== 2. LAYANAN & HARGA (11-30) =====
    ("Apa saja layanan yang tersedia?", "Kami menyediakan layanan: umum, gigi, lab, radiologi (X-Ray & USG), imunisasi, dan KIA.", "Layanan & Harga"),
    ("Berapa biaya konsultasi dokter umum?", "Biaya konsultasi dokter umum Rp 75.000.", "Layanan & Harga"),
    ("Berapa biaya konsultasi dokter gigi?", "Biaya konsultasi dokter gigi Rp 100.000.", "Layanan & Harga"),
    ("Berapa biaya lab darah lengkap?", "Biaya lab darah lengkap Rp 350.000.", "Layanan & Harga"),
    ("Berapa biaya USG?", "Biaya USG mulai dari Rp 200.000.", "Layanan & Harga"),
    ("Berapa biaya Rontgen (X-Ray)?", "Biaya Rontgen mulai dari Rp 150.000.", "Layanan & Harga"),
    ("Apakah ada harga untuk imunisasi?", "Imunisasi mulai Rp 150.000 tergantung jenis vaksin.", "Layanan & Harga"),
    ("Berapa biaya cek kehamilan?", "Biaya cek kehamilan (USG) mulai Rp 200.000.", "Layanan & Harga"),
    ("Apakah ada layanan gawat darurat?", "Kami melayani kasus gawat darurat terbatas. Untuk keadaan darurat parah, silakan ke IGD rumah sakit terdekat.", "Layanan & Harga"),
    ("Berapa biaya resep obat?", "Biaya resep obat Rp 25.000.", "Layanan & Harga"),
    ("Apakah harga sudah termasuk pajak?", "Harga yang tercantum sudah termasuk PPN 11%.", "Layanan & Harga"),
    ("Apakah ada diskon untuk paket kesehatan?", "Ya, ada paket kesehatan bulanan mulai Rp 500.000.", "Layanan & Harga"),
    ("Berapa biaya cek gula darah?", "Biaya cek gula darah Rp 75.000.", "Layanan & Harga"),
    ("Berapa biaya cek kolesterol?", "Biaya cek kolesterol Rp 100.000.", "Layanan & Harga"),
    ("Berapa biaya check-up umum?", "Check-up umum mulai Rp 500.000.", "Layanan & Harga"),
    ("Apakah ada layanan vaksinasi Covid-19?", "Ya, vaksinasi Covid-19 tersedia dengan biaya Rp 150.000/dosis.", "Layanan & Harga"),
    ("Berapa biaya sunat?", "Biaya sunat mulai Rp 500.000.", "Layanan & Harga"),
    ("Berapa biaya perawatan gigi?", "Perawatan gigi mulai Rp 150.000 (tambal).", "Layanan & Harga"),
    ("Berapa biaya cabut gigi?", "Cabut gigi sederhana Rp 150.000, complex Rp 300.000.", "Layanan & Harga"),
    ("Apakah ada layanan konsultasi psikologi?", "Ya, konsultasi psikologi Rp 200.000/sesi.", "Layanan & Harga"),
    ("Berapa biaya EKG?", "Biaya EKG Rp 150.000.", "Layanan & Harga"),
    ("Berapa biaya spirometri?", "Biaya spirometri Rp 200.000.", "Layanan & Harga"),
    ("Apakah ada layanan fisioterapi?", "Ya, fisioterapi mulai Rp 200.000/sesi.", "Layanan & Harga"),
    ("Berapa biaya suntik vitamin?", "Biaya suntik vitamin Rp 100.000.", "Layanan & Harga"),
    ("Apakah ada layanan kesehatan wanita?", "Ya, layanan KIA dan konsultasi kebidanan tersedia.", "Layanan & Harga"),
    ("Berapa biaya USG 4D?", "Biaya USG 4D Rp 500.000.", "Layanan & Harga"),
    ("Berapa biaya cek kehamilan trimester 1?", "Biaya cek kehamilan trimester 1 mulai Rp 250.000.", "Layanan & Harga"),
    ("Apakah ada layanan donor darah?", "Tidak, untuk donor darah silakan ke PMI terdekat.", "Layanan & Harga"),
    ("Berapa biaya konsultasi nutriisi?", "Konsultasi nutrisi Rp 150.000.", "Layanan & Harga"),
    ("Apakah ada layanan detoks?", "Tidak tersedia di klinik kami.", "Layanan & Harga"),

    # ===== 3. PENDAFTARAN & JADWAL (31-50) =====
    ("Bagaimana cara daftar di klinik?", "Anda bisa daftar langsung ke poliklinik atau melalui WhatsApp.", "Pendaftaran"),
    ("Apakah perlu booking sebelumnya?", "Booking tidak wajib, tapi disarankan untuk menghindari antrian.", "Pendaftaran"),
    ("Bisa daftar online?", "Belum tersedia pendaftaran online, hanya datang langsung atau via WhatsApp.", "Pendaftaran"),
    ("Berapa lama antrian biasanya?", "Antrian rata-rata 30-60 menit tergantung hari.", "Pendaftaran"),
    ("Hari apa yang paling ramai?", "Hari Senin dan Sabtu biasanya paling ramai.", "Pendaftaran"),
    ("Apakah bisa daftar ulang hari yang sama?", "Tidak, setiap kunjungan memerlukan pendaftaran baru.", "Pendaftaran"),
    ("Bagaimana sistem antrian?", "Sistem antrian berdasarkan waktu kedatangan (first come first served).", "Pendaftaran"),
    ("Apakah ada nomor antrian?", "Ya, Anda akan mendapat nomor antrian saat pendaftaran.", "Pendaftaran"),
    ("Berapa lama menunggu hasil lab?", "Hasil lab umumnya siap dalam 1-2 hari kerja.", "Pendaftaran"),
    ("Apakah bisa ambil hasil lab via WhatsApp?", "Ya, hasil lab bisa dikirim via WhatsApp.", "Pendaftaran"),
    ("Apakah perlu membawa KTP?", "Ya, wajib membawa KTP/e-KTP saat pendaftaran.", "Pendaftaran"),
    ("Apakah anak-anak perlu KTP?", "Untuk anak di bawah 17 tahun, cukup membawa Kartu Keluarga (KK).", "Pendaftaran"),
    ("Apakah bisa daftar pakai BPJS?", "Ya, bawa kartu BPJS dan surat referral dari faskes tingkat 1.", "Pendaftaran"),
    ("Berapa lama validitas rujukan BPJS?", "Rujukan BPJS berlaku 3 bulan dari tanggal diterbitkan.", "Pendaftaran"),
    ("Apakah bisa ganti dokter?", "Bisa, tapi harus mendaftar ulang.", "Pendaftaran"),
    ("Bagaimana cara membatalkan janji?", "Hubungi WhatsApp klinik minimal H-1 sebelum jadwal.", "Pendaftaran"),
    ("Apakah bisa jadwal ulang?", "Bisa, hubungi WhatsApp klinik.", "Pendaftaran"),
    ("Apakah ada layanan drive-thru?", "Tidak tersedia.", "Pendaftaran"),
    ("Apakah bisa daftar untuk orang lain?", "Bisa, tapi perlu data diri orang yang diwakili.", "Pendaftaran"),
    ("Berapa usia minimum untuk berobat?", "Tidak ada batasan usia, bayi baru lahir pun bisa berobat.", "Pendaftaran"),
    ("Apakah bisa daftar untuk hewan?", "Tidak, kami hanya melayani manusia.", "Pendaftaran"),
    ("Bagaimana jika hilang nomor antrian?", "Silakan ke loket pendaftaran untuk cetak ulang.", "Pendaftaran"),
    ("Apakah ada layanan antar结果 lab?", "Tidak, pasien harus mengambil sendiri di loket.", "Pendaftaran"),
    ("Berapa lama hasil USG keluar?", "Hasil USG bisa langsung diberikan setelah pemeriksaan.", "Pendaftaran"),
    ("Apakah ada loket khusus lansia?", "Ya, ada loket prioritas untuk lansia (60+).", "Pendaftaran"),
    ("Apakah ada loket khusus ibu hamil?", "Ya, ada loket KIA untuk ibu hamil.", "Pendaftaran"),
    ("Bagaimana jika antrian terlalu lama?", "Anda bisa menunggu di area tunggu yang nyaman atau pulang dan kembali nanti.", "Pendaftaran"),
    ("Apakah bisa konsultasi via video call?", "Belum tersedia telemedicine di klinik kami.", "Pendaftaran"),
    ("Apakah ada aplikasi antrian?", "Belum ada, hanya sistem antrian manual.", "Pendaftaran"),
    ("Bagaimana cara cek antrian?", "Cek nomor antrian di layar TV yang terpasang di ruang tunggu.", "Pendaftaran"),

    # ===== 4. PEMBAYARAN (51-70) =====
    ("Metode pembayaran apa saja?", "Tunai, transfer bank, kartu debit/kredit, dan e-wallet.", "Pembayaran"),
    ("Apakah bisa bayar pakai kartu?", "Ya, kami menerima kartu debit dan kredit.", "Pembayaran"),
    ("Apakah bisa bayar pakai e-wallet?", "Ya, kami menerima GoPay, OVO, Dana, dan LinkAja.", "Pembayaran"),
    ("Apakah bisa kredit?", "Tidak tersedia layanan kredit.", "Pembayaran"),
    ("Apakah ada cicilan?", "Tidak, pembayaran harus lunas saat itu juga.", "Pembayaran"),
    ("Apakah bisa bayar nanti?", "Tidak, pembayaran dilakukan sebelum menerima obat/perawatan.", "Pembayaran"),
    ("Berapa lama waktu pembayaran?", "Pembayaran harus lunas sebelum pasien pulang.", "Pembayaran"),
    ("Apakah ada diskon tunai?", "Tidak ada diskon khusus untuk pembayaran tunai.", "Pembayaran"),
    ("Apakah ada diskon BPJS?", "Ya, untuk pasien BPJS dikenakan tarif sesuai aturan.", "Pembayaran"),
    ("Apakah bisa claim ke asuransi?", "Ya, kami bisa buatkan surat untuk claim asuransi.", "Pembayaran"),
    ("Apakah BPJS diterima?", "Ya, kami menerima BPJS Kesehatan.", "Pembayaran"),
    ("Apakah ada tagihan elektronik?", "Ya, struk pembayaran akan dicetak.", "Pembayaran"),
    ("Bagaimana cara bayar pakai BPJS?", "Bawa kartu BPJS dan surat referral dari Faskes 1.", "Pembayaran"),
    ("Apakah bisa bayar nanti setelah pulang?", "Tidak, pembayaran dilakukan di klinik.", "Pembayaran"),
    ("Apakah ada kartu member?", "Ya, kami mengeluarkan kartu member setelah pendaftaran pertama.", "Pembayaran"),
    ("Apakah kartu member bisa dipakai berulang?", "Ya, kartu member berlaku seumur hidup.", "Pembayaran"),
    ("Bagaimana cara isi ulang paket kesehatan?", "Anda bisa beli ulang paket kesehatan di loket pendaftaran.", "Pembayaran"),
    ("Apakah ada paket bulanan?", "Ya, paket bulanan mulai Rp 500.000.", "Pembayaran"),
    ("Berapa lama masa berlaku paket?", "Masa berlaku 30 hari dari tanggal pembelian.", "Pembayaran"),
    ("Apakah bisa refund?", "Refund hanya untuk kesalahan pembayaran, hubungi admin.", "Pembayaran"),
    ("Apakah ada biaya administrasi?", "Ya, biaya administrasi Rp 10.000.", "Pembayaran"),
    ("Apakah biaya obat sudah termasuk?", "Belum, biaya obat ditagih terpisah.", "Pembayaran"),
    ("Bagaimana cara bayar untuk orang lain?", "Anda bisa membayar untuk orang lain dengan menunjukkan nama pasien.", "Pembayaran"),
    ("Apakah ada parking?", "Ya, tersedia area parkir motor dan mobil.", "Pembayaran"),
    ("Berapa biaya parkir?", "Parkir motor Rp 3.000, mobil Rp 10.000.", "Pembayaran"),
    ("Apakah parkir gratis untuk pasien?", "Tidak, parkir dibebankan kepada pasien.", "Pembayaran"),
    ("Apakah bisa bayar pakai valet?", "Tidak tersedia valet parking.", "Pembayaran"),
    ("Apakah ada ATM di klinik?", "Tidak ada ATM, tapi ada ATM di dekat klinik.", "Pembayaran"),
    ("Apakah bisa bayar pakai QRIS?", "Ya, kami menerima pembayaran QRIS.", "Pembayaran"),
    ("Bagaimana jika alat pembayaran error?", "Tetap bayar via metode lain atau coba lagi nanti.", "Pembayaran"),

    # ===== 5. OBAT & RESEP (71-85) =====
    ("Apakah ada apotek di klinik?", "Ya, ada apotek di dalam kompleks klinik.", "Obat & Resep"),
    ("Bisa beli obat di apotek?", "Ya, obat bisa dibeli langsung di apotek klinik.", "Obat & Resep"),
    ("Apakah obat bisa diambil langsung?", "Ya, obat bisa diambil setelah pembayaran.", "Obat & Resep"),
    ("Berapa lama obat jadi?", "Biasanya 15-30 menit setelah pembayaran.", "Obat & Resep"),
    ("Apakah bisa request obat tertentu?", "Bisa, tapi obat harus tersedia di apotek.", "Obat & Resep"),
    ("Apakah ada obat generik?", "Ya, tersedia obat generik dengan harga lebih murah.", "Obat & Resep"),
    ("Apakah obat bisa dikirim?", "Tidak, pasien harus mengambil sendiri.", "Obat & Resep"),
    ("Apakah resep bisa dipakai di apotek lain?", "Tidak, resep hanya valid di apotek klinik kami.", "Obat & Resep"),
    ("Bagaimana jika obat habis?", "Silakan datang ulang untuk konsultasi dan perpanjang resep.", "Obat & Resep"),
    ("Apakah ada obat bebas?", "Ya, tersedia obat-obatan bebas yang bisa dibeli tanpa resep.", "Obat & Resep"),
    ("Apakah obat mahal?", "Harga obat bervariasi, ada yang lebih murah dan ada yang lebih mahal.", "Obat & Resep"),
    ("Apakah bisa ambil obat besok?", "Tidak, obat harus diambil hari yang sama.", "Obat & Resep"),
    ("Apakah ada obat tradisional?", "Ya, tersedia beberapa obat tradisional.", "Obat & Resep"),
    ("Apakah suplemen dijual?", "Ya, tersedia suplemen kesehatan.", "Obat & Resep"),
    ("Bagaimana cara menyimpan obat?", "Ikuti instruksi dari apoteker, umumnya simpan di tempat sejuk dan kering.", "Obat & Resep"),

    # ===== 6. DOKTER & STAFF (86-100) =====
    ("Siapa dokter di klinik ini?", "Dokter kami: dr. Ahmad (Umum), dr. Siti (Gigi), dr. Budi (KIA).", "Dokter & Staff"),
    ("Apa spesifikasi dokter gigi?", "Dokter gigi spesialis konsultasi dan perawatan gigi umum.", "Dokter & Staff"),
    ("Apakah dokternya tetap?", "Ya, dokter tetap bekerja sesuai jadwal.", "Dokter & Staff"),
    ("Bagaimana jadwal dokter?", "Jadwal dokter terpasang di informasi klinik.", "Dokter & Staff"),
    ("Bisa pilih dokter?", "Bisa, tapi tergantung ketersediaan dokter hari itu.", "Dokter & Staff"),
    ("Apakah ada dokter spesialis?", "Tidak, kami hanya memiliki dokter umum dan dokter gigi.", "Dokter & Staff"),
    ("Siapa staf front desk?", "Staf front desk kami ramah dan siap membantu.", "Dokter & Staff"),
    ("Apakah staff profesional?", "Ya, semua staff terlatih dan bersertifikat.", "Dokter & Staff"),
    ("Bagaimana cara komplain?", "Silakan hubungi admin atau tulis di buku komplain yang tersedia.", "Dokter & Staff"),
    ("Apakah ada survey kepuasan?", "Ya, kami melakukan survey kepuasan pasien secara berkala.", "Dokter & Staff"),
    ("Bagaimana jika tidak puas?", "Silakan sampaikan ke admin untuk ditindaklanjuti.", "Dokter & Staff"),
    ("Apakah bisa request dokter tertentu?", "Bisa, tapi tergantung jadwal dokter tersebut.", "Dokter & Staff"),
    ("Apakah ada dokter wanita?", "Ya, tersedia dokter wanita.", "Dokter & Staff"),
    ("Apakah ada dokter anak?", "Tidak, untuk anak bisa berobat ke dokter umum kami.", "Dokter & Staff"),
    ("Bagaimana cara booking dokter spesial?", "Hubungi WhatsApp klinik untuk booking dokter spesial.", "Dokter & Staff"),
]

# Buat workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "FAQ Klinik"

# Header style
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
headers = ["No", "Pertanyaan", "Jawaban", "Kategori"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Write data
for idx, (q, a, cat) in enumerate(faq_data, 1):
    ws.cell(row=idx+1, column=1, value=idx).border = border
    ws.cell(row=idx+1, column=2, value=q).border = border
    ws.cell(row=idx+1, column=3, value=a).border = border
    ws.cell(row=idx+1, column=4, value=cat).border = border
    
    # Wrap text for question and answer
    ws.cell(row=idx+1, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=idx+1, column=3).alignment = Alignment(wrap_text=True, vertical="top")

# Set column widths
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 18

# Save
output_path = "/media/ahmad/84a8377e-0bbf-4a05-bc83-75f57016cb6c/bisnis/ai_agent/chatbot/data/klinik_faq_100.xlsx"
wb.save(output_path)
print(f"✅ Spreadsheet saved: {output_path}")
print(f"📊 Total FAQ: {len(faq_data)}")
